from flask import Flask, request, jsonify
from flask_cors import CORS
import os
import shutil
import re
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
CORS(app)

BASE_DIR = "data"
UPLOAD_DIR = "uploads"

os.makedirs(BASE_DIR, exist_ok=True)
os.makedirs(UPLOAD_DIR, exist_ok=True)


# ---------------- BASIC ----------------
@app.route("/")
def home():
    return "Backend running"


def safe_name(name):
    return name.replace(" ", "_")
# ---------------- ADMIN: LIST EXAMS ----------------
@app.route("/admin/exams", methods=["GET"])
def admin_list_exams():
    exams = []

    for folder in os.listdir(BASE_DIR):
        exam_dir = os.path.join(BASE_DIR, folder)
        scheme_path = os.path.join(exam_dir, "marking_scheme.xlsx")

        if not os.path.isdir(exam_dir) or not os.path.exists(scheme_path):
            continue

        wb = load_workbook(scheme_path)
        ws = wb.active

        scheme = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            scheme[row[0]] = row[1]

        exams.append({
            "exam_name": scheme.get("Exam Name"),
            "correct": scheme.get("Correct"),
            "wrong": scheme.get("Wrong"),
            "na": scheme.get("NA")
        })

    return jsonify(exams)


# ---------------- ADMIN: CREATE EXAM ----------------
@app.route("/admin/create-exam", methods=["POST"])
def create_exam():
    data = request.json

    exam_name = data.get("exam_name")
    correct = data.get("correct")
    wrong = data.get("wrong")
    na = data.get("na")

    if not exam_name:
        return jsonify({"error": "exam_name required"}), 400

    exam_dir = os.path.join(BASE_DIR, safe_name(exam_name))
    os.makedirs(exam_dir, exist_ok=True)

    # ---- marking_scheme.xlsx ----
    scheme_path = os.path.join(exam_dir, "marking_scheme.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "scheme"
    ws.append(["Key", "Value"])
    ws.append(["Exam Name", exam_name])
    ws.append(["Correct", correct])
    ws.append(["Wrong", wrong])
    ws.append(["NA", na])
    wb.save(scheme_path)

    # ---- responses.xlsx ----
    response_path = os.path.join(exam_dir, "responses.xlsx")
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "responses"
    ws2.append([
        "Name", "Roll", "Category", "Gender", "State",
        "Total Marks"
    ])
    wb2.save(response_path)

    return jsonify({"status": "success", "exam": exam_name})

# ---------------- ADMIN: DELETE EXAM ----------------
@app.route("/admin/delete-exam", methods=["POST"])
def delete_exam():
    data = request.json
    if not data or "exam_name" not in data:
        return jsonify({"error": "exam_name required"}), 400

    exam_name = data["exam_name"]
    exam_dir = os.path.join(BASE_DIR, safe_name(exam_name))

    if not os.path.exists(exam_dir):
        return jsonify({"error": "exam not found"}), 404

    shutil.rmtree(exam_dir)
    return jsonify({"status": "deleted", "exam": exam_name})

# ---------------- READ MARKING SCHEME ----------------
def read_marking_scheme(exam_name):
    path = os.path.join(BASE_DIR, safe_name(exam_name), "marking_scheme.xlsx")
    wb = load_workbook(path)
    ws = wb.active

    scheme = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        scheme[row[0]] = row[1]
    return scheme


# ---------------- PARSE SECTION-WISE RESPONSE ----------------
def parse_response_sectionwise(html_path, scheme):
    section_map = {}
    section_order = []
    current_section = None

    with open(html_path, "r", encoding="utf-8", errors="ignore") as f:
        soup = BeautifulSoup(f, "lxml")

    for el in soup.find_all("div"):
        if el.get("class") == ["section-lbl"]:
            current_section = el.get_text(strip=True)
            if current_section not in section_map:
                section_map[current_section] = {
                    "correct": 0,
                    "wrong": 0,
                    "na": 0
                }
                section_order.append(current_section)

        if el.get("class") == ["question-pnl"] and current_section:
            chosen = None
            correct_option = None

            chosen_row = el.find("td", string=re.compile("Chosen Option"))
            if chosen_row:
                v = chosen_row.find_next_sibling("td").get_text(strip=True)
                if v != "--":
                    chosen = int(v)

            right_td = el.find("td", class_="rightAns")
            if right_td:
                m = re.search(r"(\d)\.", right_td.get_text())
                if m:
                    correct_option = int(m.group(1))

            if chosen is None:
                section_map[current_section]["na"] += 1
            elif chosen == correct_option:
                section_map[current_section]["correct"] += 1
            else:
                section_map[current_section]["wrong"] += 1

    subject_marks = []
    subject_stats = []
    total_marks = 0

    for sec in section_order:
        c = section_map[sec]["correct"]
        w = section_map[sec]["wrong"]
        n = section_map[sec]["na"]
        attempt = c + w

        marks = (
            c * scheme["Correct"]
            + w * scheme["Wrong"]
            + n * scheme["NA"]
        )

        subject_marks.append(marks)
        subject_stats.append({
            "attempt": attempt,
            "right": c,
            "wrong": w,
            "na": n
        })

        total_marks += marks

    return total_marks, subject_marks, subject_stats


# ---------------- SAVE RESULT (BACKWARD COMPATIBLE) ----------------
def save_user_result(exam_name, base_data, subject_marks, subject_stats):
    path = os.path.join(BASE_DIR, safe_name(exam_name), "responses.xlsx")
    wb = load_workbook(path)
    ws = wb.active

    headers = [c.value for c in ws[1]]

    # ---- ensure Subject N (old behavior) ----
    for i in range(len(subject_marks)):
        col_name = f"Subject {i+1}"
        if col_name not in headers:
            ws.cell(row=1, column=len(headers) + 1, value=col_name)
            headers.append(col_name)

    # ---- append base + marks (unchanged behavior) ----
    ws.append(base_data + subject_marks)
    current_row = ws.max_row

    # ---- OPTIONAL stats (new behavior) ----
    for i, stats in enumerate(subject_stats):
        base = f"S{i+1}"
        mapping = {
            f"{base}_Attempt": stats["attempt"],
            f"{base}_R": stats["right"],
            f"{base}_W": stats["wrong"],
            f"{base}_NA": stats["na"]
        }

        for h, v in mapping.items():
            if h not in headers:
                ws.cell(row=1, column=len(headers) + 1, value=h)
                headers.append(h)
            ws.cell(row=current_row, column=headers.index(h) + 1, value=v)

    wb.save(path)


# ---------------- EVALUATE ----------------
@app.route("/evaluate", methods=["POST"])
def evaluate_exam():
    exam_name = request.form.get("exam_name")
    name = request.form.get("name")
    roll = request.form.get("roll")
    category = request.form.get("category")
    gender = request.form.get("gender")
    state = request.form.get("state")
    file = request.files.get("file")

    if not all([exam_name, name, roll, category, gender, state, file]):
        return jsonify({"error": "Missing fields"}), 400

    upload_path = os.path.join(UPLOAD_DIR, file.filename)
    file.save(upload_path)

    scheme = read_marking_scheme(exam_name)

    total_marks, subject_marks, subject_stats = parse_response_sectionwise(
        upload_path, scheme
    )

    save_user_result(
        exam_name,
        [name, roll, category, gender, state, total_marks],
        subject_marks,
        subject_stats
    )

    return jsonify({"status": "saved"})


# ---------------- RESULT (SMART FALLBACK) ----------------
@app.route("/result")
def get_result():
    exam = request.args.get("exam")
    roll = request.args.get("roll")

    if not exam or not roll:
        return jsonify({"error": "Missing exam or roll"}), 400

    path = os.path.join(BASE_DIR, exam, "responses.xlsx")
    if not os.path.exists(path):
        return jsonify({"error": "Result file not found"}), 404

    wb = load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    row = None
    for r in ws.iter_rows(min_row=2, values_only=True):
        if str(r[col["Roll"]]) == str(roll):
            row = r
            break

    if not row:
        return jsonify({"error": "Candidate not found"}), 404

    subjects = []
    overall = {"attempt": 0, "right": 0, "wrong": 0, "na": 0, "marks": 0}

    i = 1
    while f"Subject {i}" in headers:
        def safe(h):
            return row[col[h]] if h in col and row[col[h]] is not None else "-"

        marks = row[col[f"Subject {i}"]]

        subjects.append({
            "name": f"Subject {i}",
            "attempt": safe(f"S{i}_Attempt"),
            "right": safe(f"S{i}_R"),
            "wrong": safe(f"S{i}_W"),
            "na": safe(f"S{i}_NA"),
            "marks": marks
        })

        if isinstance(safe(f"S{i}_Attempt"), int):
            overall["attempt"] += safe(f"S{i}_Attempt")
            overall["right"] += safe(f"S{i}_R")
            overall["wrong"] += safe(f"S{i}_W")
            overall["na"] += safe(f"S{i}_NA")

        overall["marks"] += marks
        i += 1

    return jsonify({
        "exam": exam,
        "candidate": {
            "name": row[col["Name"]],
            "roll": row[col["Roll"]],
            "category": row[col["Category"]],
            "gender": row[col["Gender"]],
            "state": row[col["State"]]
        },
        "subjects": subjects,
        "overall": overall
    })


# ---------------- START ----------------
if __name__ == "__main__":
    app.run(debug=True)
